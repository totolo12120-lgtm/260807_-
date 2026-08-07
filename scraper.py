import requests
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
from datetime import datetime, timedelta
from urllib.parse import urljoin
import urllib3
import os

# SSL 경고 비활성화
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
}

KEYWORDS = ["청년", "복지", "지원", "인공지능", "AI", "교육"]

def get_cutoff_date():
    return datetime.now().date() - timedelta(days=30)

def parse_date(date_str):
    if not date_str:
        return None
    clean_str = re.sub(r'[^\d]', '', date_str)
    if len(clean_str) >= 8:
        try:
            return datetime.strptime(clean_str[:8], '%Y%m%d').date()
        except ValueError:
            pass
    return None

def clean_text(text):
    if not text:
        return ""
    return re.sub(r'\s+', ' ', text).strip()

def get_matched_keywords(text):
    if not text:
        return ""
    text_upper = text.upper()
    matched = []
    for kw in KEYWORDS:
        if kw.upper() in text_upper:
            matched.append(kw)
    return ", ".join(matched)

def check_keyword(title):
    return len(get_matched_keywords(title)) > 0

# 1. 행정안전부 (MOIS)
def scrape_mois(cutoff_date):
    print("[1/3] 행정안전부 공고 수집 중...")
    results = []
    base_url = 'https://www.mois.go.kr/frt/bbs/type013/commonSelectBoardList.do'
    page = 1
    stop = False
    
    while page <= 10 and not stop:
        url = f"{base_url}?bbsId=BBSMSTR_000000000006&pageIndex={page}"
        try:
            res = requests.get(url, headers=HEADERS, verify=False, timeout=10)
            res.encoding = res.apparent_encoding or 'utf-8'
            soup = BeautifulSoup(res.text, 'html.parser')
            table = soup.find('table')
            if not table:
                break
            rows = table.find_all('tr')[1:]
            if not rows:
                break
            for tr in rows:
                tds = tr.find_all('td')
                if len(tds) < 5:
                    continue
                title_td = tds[1]
                a_tag = title_td.find('a')
                if not a_tag:
                    continue
                title = clean_text(title_td.get_text())
                reg_date_str = tds[4].get_text(strip=True)
                reg_date = parse_date(reg_date_str)
                
                if reg_date and reg_date < cutoff_date:
                    stop = True
                    break
                    
                href = a_tag.get('href', '')
                link = urljoin('https://www.mois.go.kr', href)
                
                matched_kws = get_matched_keywords(title)
                if matched_kws:
                    results.append({
                        '지자체명': '행정안전부',
                        '공고제목': title,
                        '매칭키워드': matched_kws,
                        '등록일': str(reg_date) if reg_date else reg_date_str,
                        '마감일': '상시/정보없음',
                        '링크': link
                    })
        except Exception as e:
            print(f"행정안전부 수집 중 오류: {e}")
            break
        page += 1
        
    print(f"-> 행정안전부 완료: {len(results)}건 수집됨")
    return results

# 2. 중소벤처기업부 (MSS)
def scrape_mss(cutoff_date):
    print("[2/3] 중소벤처기업부 사업공고 수집 중...")
    results = []
    base_url = 'https://www.mss.go.kr/site/smba/ex/bbs/List.do'
    page = 1
    stop = False
    
    while page <= 10 and not stop:
        url = f"{base_url}?cbIdx=310&pageIndex={page}"
        try:
            res = requests.get(url, headers=HEADERS, verify=False, timeout=10)
            res.encoding = res.apparent_encoding or 'utf-8'
            soup = BeautifulSoup(res.text, 'html.parser')
            table = soup.find('table')
            if not table:
                break
            rows = table.find_all('tr')[1:]
            if not rows:
                break
            for tr in rows:
                tds = tr.find_all('td')
                if len(tds) < 4:
                    continue
                
                onclick = tr.get('onclick', '')
                bc_match = re.search(r"doBbsFView\([^,]+,\s*'(\d+)'", onclick)
                if not bc_match:
                    file_span = tr.find('span', class_='single-file')
                    if file_span:
                        data_href = file_span.get('data-href', '')
                        bc_match = re.search(r"bcIdx=(\d+)", data_href)
                        
                bc_idx = bc_match.group(1) if bc_match else ''
                link = f"https://www.mss.go.kr/site/smba/ex/bbs/View.do?cbIdx=310&bcIdx={bc_idx}" if bc_idx else 'https://www.mss.go.kr/site/smba/ex/bbs/List.do?cbIdx=310'
                
                title_td = tds[1]
                a_tag = title_td.find('a')
                raw_title = clean_text(title_td.get_text(" "))
                clean_t = clean_text(a_tag.get_text()) if a_tag else raw_title
                
                reg_date_str = tds[3].get_text(strip=True)
                reg_date = parse_date(reg_date_str)
                
                if reg_date and reg_date < cutoff_date:
                    stop = True
                    break
                    
                deadline = '상시/정보없음'
                dl_match = re.search(r'신청기간\s*[:\s]*[\d\.\-~]+\s*~\s*([\d\.\-]+)', raw_title)
                if dl_match:
                    parsed_dl = parse_date(dl_match.group(1))
                    deadline = str(parsed_dl) if parsed_dl else dl_match.group(1)
                    
                matched_kws = get_matched_keywords(raw_title)
                if matched_kws:
                    results.append({
                        '지자체명': '중소벤처기업부',
                        '공고제목': clean_t,
                        '매칭키워드': matched_kws,
                        '등록일': str(reg_date) if reg_date else reg_date_str,
                        '마감일': deadline,
                        '링크': link
                    })
        except Exception as e:
            print(f"중소벤처기업부 수집 중 오류: {e}")
            break
        page += 1
        
    print(f"-> 중소벤처기업부 완료: {len(results)}건 수집됨")
    return results

# 3. 고용노동부 (MOEL)
def scrape_moel(cutoff_date):
    print("[3/3] 고용노동부 공고 수집 중...")
    results = []
    base_url = 'https://www.moel.go.kr/news/notice/noticeList.do'
    page = 1
    stop = False
    
    while page <= 10 and not stop:
        url = f"{base_url}?pageIndex={page}"
        try:
            res = requests.get(url, headers=HEADERS, verify=False, timeout=10)
            res.encoding = res.apparent_encoding or 'utf-8'
            soup = BeautifulSoup(res.text, 'html.parser')
            table = soup.find('table')
            if not table:
                break
            rows = table.find_all('tr')[1:]
            if not rows:
                break
            for tr in rows:
                tds = tr.find_all('td')
                if len(tds) < 5:
                    continue
                title_td = tds[1]
                a_tag = title_td.find('a')
                if not a_tag:
                    continue
                title = clean_text(a_tag.get_text())
                reg_date_str = tds[4].get_text(strip=True)
                reg_date = parse_date(reg_date_str)
                
                if reg_date and reg_date < cutoff_date:
                    stop = True
                    break
                    
                href = a_tag.get('href', '')
                link = urljoin('https://www.moel.go.kr', href)
                
                matched_kws = get_matched_keywords(title)
                if matched_kws:
                    results.append({
                        '지자체명': '고용노동부',
                        '공고제목': title,
                        '매칭키워드': matched_kws,
                        '등록일': str(reg_date) if reg_date else reg_date_str,
                        '마감일': '상시/정보없음',
                        '링크': link
                    })
        except Exception as e:
            print(f"고용노동부 수집 중 오류: {e}")
            break
        page += 1
        
    print(f"-> 고용노동부 완료: {len(results)}건 수집됨")
    return results

def save_to_excel(mois_data, mss_data, moel_data):
    today_str = datetime.now().strftime('%Y%m%d')
    base_file_name = f"타기관벤치마킹_{today_str}.xlsx"
    file_name = base_file_name
    
    df_mois = pd.DataFrame(mois_data)
    df_mss = pd.DataFrame(mss_data)
    df_moel = pd.DataFrame(moel_data)
    
    cols = ['지자체명', '공고제목', '매칭키워드', '등록일', '마감일', '링크']
    for df in [df_mois, df_mss, df_moel]:
        for col in cols:
            if col not in df.columns:
                df[col] = []
                
    df_total = pd.concat([df_mois, df_mss, df_moel], ignore_index=True)
    if not df_total.empty:
        df_total = df_total.sort_values(by='등록일', ascending=False)
        
    try:
        writer = pd.ExcelWriter(file_name, engine='openpyxl')
        df_total.to_excel(writer, sheet_name='통합비교표', index=False)
        df_mois.to_excel(writer, sheet_name='행정안전부', index=False)
        df_mss.to_excel(writer, sheet_name='중소벤처기업부', index=False)
        df_moel.to_excel(writer, sheet_name='고용노동부', index=False)
        writer.close()
    except PermissionError:
        file_name = f"타기관벤치마킹_{today_str}_new.xlsx"
        print(f"\n[주의] {base_file_name} 파일이 열려 있어 {file_name} 로 저장합니다.")
        writer = pd.ExcelWriter(file_name, engine='openpyxl')
        df_total.to_excel(writer, sheet_name='통합비교표', index=False)
        df_mois.to_excel(writer, sheet_name='행정안전부', index=False)
        df_mss.to_excel(writer, sheet_name='중소벤처기업부', index=False)
        df_moel.to_excel(writer, sheet_name='고용노동부', index=False)
        writer.close()
        
    # Excel Formatting
    wb = openpyxl.load_workbook(file_name)
    
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.views.sheetView[0].showGridLines = True
        
        # Header formatting
        for col_num in range(1, 7):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            
        # Rows formatting
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 22
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.font = data_font
                cell.border = thin_border
                
                # Column alignments
                if col in [1, 3, 4, 5]:  # 지자체명, 매칭키워드, 등록일, 마감일
                    cell.alignment = center_align
                else:  # 공고제목, 링크
                    cell.alignment = left_align
                    
        # Adjust column widths
        ws.column_dimensions['A'].width = 16  # 지자체명
        ws.column_dimensions['B'].width = 65  # 공고제목
        ws.column_dimensions['C'].width = 20  # 매칭키워드
        ws.column_dimensions['D'].width = 14  # 등록일
        ws.column_dimensions['E'].width = 16  # 마감일
        ws.column_dimensions['F'].width = 60  # 링크

    wb.save(file_name)
    print(f"\n성공적으로 엑셀 파일이 저장되었습니다: {file_name}")
    return file_name

def main():
    cutoff_date = get_cutoff_date()
    print(f"=== 청년·복지·지원·AI·교육 사업공고 크롤링 시작 ===")
    print(f"수집 기준일: {cutoff_date} 이후 (최근 30일)\n")
    
    mois_data = scrape_mois(cutoff_date)
    mss_data = scrape_mss(cutoff_date)
    moel_data = scrape_moel(cutoff_date)
    
    save_to_excel(mois_data, mss_data, moel_data)

if __name__ == '__main__':
    main()
