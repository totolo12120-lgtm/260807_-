import sys
import os
import re
import webbrowser
import threading
from datetime import datetime, timedelta
from urllib.parse import urljoin
import urllib3

import requests
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import tkinter as tk
from tkinter import ttk, messagebox

# SSL 경고 비활성화
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
}

KEYWORDS = ["청년", "복지", "지원", "인공지능", "AI", "교육"]

def get_cutoff_date():
    return datetime.now().date() - timedelta(days=30)

def parse_date(date_str):
    if not date_str:
        return None
    clean_str = re.sub(r'[^\d]', '', date_str)
    if len(clean_str) >= 8:
        try:
            return datetime.strptime(clean_str[:8], '%Y%m%d').date()
        except ValueError:
            pass
    return None

def clean_text(text):
    if not text:
        return ""
    return re.sub(r'\s+', ' ', text).strip()

def check_keyword(title):
    if not title:
        return False
    title_upper = title.upper()
    for kw in KEYWORDS:
        if kw.upper() in title_upper:
            return True
    return False

def extract_core_keyword(title):
    if not title:
        return "일반 공고"
    
    clean = re.sub(r'\[.*?\]|「.*?」|\(.*?\)', ' ', title)
    clean = re.sub(r'\s+', ' ', clean).strip()
    
    if re.search(r'합격자\s*공고', clean):
        return "합격자 공고"
    if re.search(r'참여기업\s*(?:\d+차\s*)?모집\s*공고', clean):
        return "참여기업 모집공고"
    if re.search(r'창업기업\s*모집\s*공고', clean):
        return "창업기업 모집공고"
    if re.search(r'경력(?:경쟁)?채용시험\s*공고', clean):
        return "경력채용시험 공고"
    if re.search(r'공개모집|공모집', clean):
        if '이사장' in title:
            return "이사장 공개모집"
        if '직위' in title:
            return "직위 공개모집"
        return "공개모집"
    if re.search(r'선정\s*결과\s*공고', clean):
        if '보조사업자' in title:
            return "보조사업자 선정 결과"
        return "선정 결과 공고"
    if re.search(r'기술능력평가\s*결과\s*알림', clean):
        return "기술능력평가 결과 알림"
    if re.search(r'수상작\s*선정\s*결과', clean):
        return "공모전 수상작 선정 결과"
    if re.search(r'행정예고', clean):
        return "규정 제정 행정예고"
    if re.search(r'포상\s*후보자\s*공모', clean):
        return "포상 후보자 공모"
    if re.search(r'신규\s*지정\s*공모', clean):
        return "신규 지정 공모"
    if re.search(r'교육기관\s*선정', clean):
        return "교육기관 선정 공고"
    if re.search(r'창업경진대회', clean):
        return "창업경진대회 공고"
    if re.search(r'지원사업\s*공고', clean):
        return "지원사업 공고"
    if re.search(r'지원계획\s*공고', clean):
        return "지원계획 공고"
        
    match = re.search(r'([가-힣A-Za-z0-9\s]{2,15}(?:모집|공고|공모|결과|알림|예고))$', clean)
    if match:
        return match.group(1).strip()
        
    return "사업공고"

# Scraper Logic
def scrape_all_sites(log_callback=None):
    cutoff_date = get_cutoff_date()
    all_results = {'mois': [], 'mss': [], 'moel': []}

    # 1. MOIS
    if log_callback: log_callback("[1/3] 행정안전부 공고 수집 중...")
    base_url = 'https://www.mois.go.kr/frt/bbs/type013/commonSelectBoardList.do'
    page = 1
    stop = False
    while page <= 10 and not stop:
        url = f"{base_url}?bbsId=BBSMSTR_000000000006&pageIndex={page}"
        try:
            res = requests.get(url, headers=HEADERS, verify=False, timeout=10)
            res.encoding = res.apparent_encoding or 'utf-8'
            soup = BeautifulSoup(res.text, 'html.parser')
            table = soup.find('table')
            if not table: break
            rows = table.find_all('tr')[1:]
            if not rows: break
            for tr in rows:
                tds = tr.find_all('td')
                if len(tds) < 5: continue
                title_td = tds[1]
                a_tag = title_td.find('a')
                if not a_tag: continue
                title = clean_text(title_td.get_text())
                reg_date_str = tds[4].get_text(strip=True)
                reg_date = parse_date(reg_date_str)
                if reg_date and reg_date < cutoff_date:
                    stop = True
                    break
                href = a_tag.get('href', '')
                link = urljoin('https://www.mois.go.kr', href)
                if check_keyword(title):
                    all_results['mois'].append({
                        '지자체명': '행정안전부',
                        '공고제목': title,
                        '매칭키워드': extract_core_keyword(title),
                        '등록일': str(reg_date) if reg_date else reg_date_str,
                        '마감일': '상시/정보없음',
                        '링크': link
                    })
        except Exception as e:
            if log_callback: log_callback(f"행정안전부 오류: {e}")
            break
        page += 1
    if log_callback: log_callback(f"-> 행정안전부 수집 완료: {len(all_results['mois'])}건")

    # 2. MSS
    if log_callback: log_callback("[2/3] 중소벤처기업부 사업공고 수집 중...")
    base_url = 'https://www.mss.go.kr/site/smba/ex/bbs/List.do'
    page = 1
    stop = False
    while page <= 10 and not stop:
        url = f"{base_url}?cbIdx=310&pageIndex={page}"
        try:
            res = requests.get(url, headers=HEADERS, verify=False, timeout=10)
            res.encoding = res.apparent_encoding or 'utf-8'
            soup = BeautifulSoup(res.text, 'html.parser')
            table = soup.find('table')
            if not table: break
            rows = table.find_all('tr')[1:]
            if not rows: break
            for tr in rows:
                tds = tr.find_all('td')
                if len(tds) < 4: continue
                onclick = tr.get('onclick', '')
                bc_match = re.search(r"doBbsFView\([^,]+,\s*'(\d+)'", onclick)
                if not bc_match:
                    file_span = tr.find('span', class_='single-file')
                    if file_span:
                        data_href = file_span.get('data-href', '')
                        bc_match = re.search(r"bcIdx=(\d+)", data_href)
                bc_idx = bc_match.group(1) if bc_match else ''
                link = f"https://www.mss.go.kr/site/smba/ex/bbs/View.do?cbIdx=310&bcIdx={bc_idx}" if bc_idx else 'https://www.mss.go.kr/site/smba/ex/bbs/List.do?cbIdx=310'
                title_td = tds[1]
                a_tag = title_td.find('a')
                raw_title = clean_text(title_td.get_text(" "))
                clean_t = clean_text(a_tag.get_text()) if a_tag else raw_title
                reg_date_str = tds[3].get_text(strip=True)
                reg_date = parse_date(reg_date_str)
                if reg_date and reg_date < cutoff_date:
                    stop = True
                    break
                deadline = '상시/정보없음'
                dl_match = re.search(r'신청기간\s*[:\s]*[\d\.\-~]+\s*~\s*([\d\.\-]+)', raw_title)
                if dl_match:
                    parsed_dl = parse_date(dl_match.group(1))
                    deadline = str(parsed_dl) if parsed_dl else dl_match.group(1)
                if check_keyword(clean_t) or check_keyword(raw_title):
                    all_results['mss'].append({
                        '지자체명': '중소벤처기업부',
                        '공고제목': clean_t,
                        '매칭키워드': extract_core_keyword(clean_t),
                        '등록일': str(reg_date) if reg_date else reg_date_str,
                        '마감일': deadline,
                        '링크': link
                    })
        except Exception as e:
            if log_callback: log_callback(f"중소벤처기업부 오류: {e}")
            break
        page += 1
    if log_callback: log_callback(f"-> 중소벤처기업부 수집 완료: {len(all_results['mss'])}건")

    # 3. MOEL
    if log_callback: log_callback("[3/3] 고용노동부 공고 수집 중...")
    base_url = 'https://www.moel.go.kr/news/notice/noticeList.do'
    page = 1
    stop = False
    while page <= 10 and not stop:
        url = f"{base_url}?pageIndex={page}"
        try:
            res = requests.get(url, headers=HEADERS, verify=False, timeout=10)
            res.encoding = res.apparent_encoding or 'utf-8'
            soup = BeautifulSoup(res.text, 'html.parser')
            table = soup.find('table')
            if not table: break
            rows = table.find_all('tr')[1:]
            if not rows: break
            for tr in rows:
                tds = tr.find_all('td')
                if len(tds) < 5: continue
                title_td = tds[1]
                a_tag = title_td.find('a')
                if not a_tag: continue
                title = clean_text(a_tag.get_text())
                reg_date_str = tds[4].get_text(strip=True)
                reg_date = parse_date(reg_date_str)
                if reg_date and reg_date < cutoff_date:
                    stop = True
                    break
                href = a_tag.get('href', '')
                link = urljoin('https://www.moel.go.kr', href)
                if check_keyword(title):
                    all_results['moel'].append({
                        '지자체명': '고용노동부',
                        '공고제목': title,
                        '매칭키워드': extract_core_keyword(title),
                        '등록일': str(reg_date) if reg_date else reg_date_str,
                        '마감일': '상시/정보없음',
                        '링크': link
                    })
        except Exception as e:
            if log_callback: log_callback(f"고용노동부 오류: {e}")
            break
        page += 1
    if log_callback: log_callback(f"-> 고용노동부 수집 완료: {len(all_results['moel'])}건")

    return all_results

def export_to_excel(all_results):
    today_str = datetime.now().strftime('%Y%m%d')
    file_name = f"타기관벤치마킹_{today_str}_핵심키워드.xlsx"
    
    df_mois = pd.DataFrame(all_results['mois'])
    df_mss = pd.DataFrame(all_results['mss'])
    df_moel = pd.DataFrame(all_results['moel'])
    
    cols = ['지자체명', '공고제목', '매칭키워드', '등록일', '마감일', '링크']
    for df in [df_mois, df_mss, df_moel]:
        for col in cols:
            if col not in df.columns:
                df[col] = []
                
    df_total = pd.concat([df_mois, df_mss, df_moel], ignore_index=True)
    if not df_total.empty:
        df_total = df_total.sort_values(by='등록일', ascending=False)

    try:
        writer = pd.ExcelWriter(file_name, engine='openpyxl')
        df_total.to_excel(writer, sheet_name='통합비교표', index=False)
        df_mois.to_excel(writer, sheet_name='행정안전부', index=False)
        df_mss.to_excel(writer, sheet_name='중소벤처기업부', index=False)
        df_moel.to_excel(writer, sheet_name='고용노동부', index=False)
        writer.close()
    except PermissionError:
        file_name = f"타기관벤치마킹_{today_str}_핵심키워드_v2.xlsx"
        writer = pd.ExcelWriter(file_name, engine='openpyxl')
        df_total.to_excel(writer, sheet_name='통합비교표', index=False)
        df_mois.to_excel(writer, sheet_name='행정안전부', index=False)
        df_mss.to_excel(writer, sheet_name='중소벤처기업부', index=False)
        df_moel.to_excel(writer, sheet_name='고용노동부', index=False)
        writer.close()
        
    # Excel Formatting
    wb = openpyxl.load_workbook(file_name)
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.views.sheetView[0].showGridLines = True
        for col_num in range(1, 7):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 22
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.font = data_font
                cell.border = thin_border
                if col in [1, 3, 4, 5]:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 65
        ws.column_dimensions['C'].width = 24
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 60

    wb.save(file_name)
    return file_name, df_total

# GUI Application
class GovernmentNoticeApp:
    def __init__(self, root):
        self.root = root
        self.root.title("공공기관 사업공고 자동 수집기 (핵심키워드 분석)")
        self.root.geometry("1100 excavation".split()[0] if False else "1150x700")
        self.root.minsize(950, 600)
        
        self.df_data = pd.DataFrame()
        self.current_excel_path = ""
        
        self.setup_styles()
        self.create_widgets()
        self.load_latest_excel_if_exists()
        
    def setup_styles(self):
        self.style = ttk.Style()
        self.style.theme_use('clam')
        
        # Colors
        self.bg_color = "#F5F7FA"
        self.header_bg = "#1F497D"
        self.accent_color = "#2563EB"
        
        self.root.configure(bg=self.bg_color)
        
        self.style.configure("Treeview.Heading", font=("맑은 고딕", 10, "bold"), background="#1F497D", foreground="white", relief="flat")
        self.style.configure("Treeview", font=("맑은 고딕", 9), rowheight=28)
        self.style.map("Treeview", background=[("selected", "#3B82F6")], foreground=[("selected", "white")])
        
    def create_widgets(self):
        # 1. Header Frame
        header_frame = tk.Frame(self.root, bg="#1F497D", height=65)
        header_frame.pack(fill="x", side="top")
        
        title_label = tk.Label(
            header_frame, 
            text="🏛️ 공공기관 청년·복지·AI 사업공고 자동 수집기", 
            font=("맑은 고딕", 16, "bold"), 
            bg="#1F497D", 
            fg="white",
            anchor="w"
        )
        title_label.pack(side="left", padx=20, pady=15)
        
        subtitle_label = tk.Label(
            header_frame, 
            text="행정안전부 | 중소벤처기업부 | 고용노동부", 
            font=("맑은 고딕", 10), 
            bg="#1F497D", 
            fg="#DBEAFE",
            anchor="e"
        )
        subtitle_label.pack(side="right", padx=20, pady=20)
        
        # 2. Control Bar Frame
        control_frame = tk.Frame(self.root, bg=self.bg_color, padx=15, pady=10)
        control_frame.pack(fill="x")
        
        self.btn_collect = tk.Button(
            control_frame, 
            text="🚀 공고 실시간 수집 시작", 
            font=("맑은 고딕", 10, "bold"), 
            bg="#2563EB", 
            fg="white", 
            activebackground="#1D4ED8", 
            activeforeground="white",
            relief="flat", 
            padx=15, 
            pady=6,
            command=self.start_scraping_thread
        )
        self.btn_collect.pack(side="left", padx=(0, 10))
        
        self.btn_open_excel = tk.Button(
            control_frame, 
            text="📂 엑셀 파일 열기", 
            font=("맑은 고딕", 10), 
            bg="#10B981", 
            fg="white", 
            activebackground="#059669", 
            activeforeground="white",
            relief="flat", 
            padx=15, 
            pady=6,
            command=self.open_excel_file
        )
        self.btn_open_excel.pack(side="left", padx=(0, 15))
        
        # Search Box
        tk.Label(control_frame, text="검색어:", font=("맑은 고딕", 10), bg=self.bg_color).pack(side="left", padx=(10, 5))
        self.search_var = tk.StringVar()
        self.search_var.trace("w", self.filter_table)
        self.search_entry = tk.Entry(control_frame, textvariable=self.search_var, font=("맑은 고딕", 10), width=20)
        self.search_entry.pack(side="left")
        
        # Status Label
        self.lbl_status = tk.Label(control_frame, text="상태: 준비 완료", font=("맑은 고딕", 10, "bold"), fg="#4B5563", bg=self.bg_color)
        self.lbl_status.pack(side="right", padx=10)
        
        # 3. Main Data Table Frame
        table_frame = tk.Frame(self.root, bg=self.bg_color, padx=15, pady=5)
        table_frame.pack(fill="both", expand=True)
        
        cols = ("지자체명", "공고제목", "매칭키워드", "등록일", "마감일", "링크")
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings", selectmode="browse")
        
        # Column headings and widths
        self.tree.heading("지자체명", text="지자체명")
        self.tree.heading("공고제목", text="공고제목")
        self.tree.heading("매칭키워드", text="매칭키워드(핵심주제)")
        self.tree.heading("등록일", text="등록일")
        self.tree.heading("마감일", text="마감일")
        self.tree.heading("링크", text="공고 바로가기 링크")
        
        self.tree.column("지자체명", width=120, anchor="center")
        self.tree.column("공고제목", width=420, anchor="w")
        self.tree.column("매칭키워드", width=150, anchor="center")
        self.tree.column("등록일", width=100, anchor="center")
        self.tree.column("마감일", width=120, anchor="center")
        self.tree.column("링크", width=220, anchor="w")
        
        # Scrollbars
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)
        
        # Bind double click on row
        self.tree.bind("<Double-1>", self.on_row_double_click)
        
        # 4. Footer Info Bar
        footer_frame = tk.Frame(self.root, bg="#E5E7EB", height=30)
        footer_frame.pack(fill="x", side="bottom")
        
        self.lbl_footer = tk.Label(
            footer_frame, 
            text="💡 목록에서 항목을 더블클릭하면 해당 웹사이트 공고 페이지로 바로 이동합니다.", 
            font=("맑은 고딕", 9), 
            bg="#E5E7EB", 
            fg="#374151"
        )
        self.lbl_footer.pack(side="left", padx=15, pady=5)
        
    def log(self, text):
        self.lbl_status.config(text=f"상태: {text}")
        self.root.update_idletasks()

    def start_scraping_thread(self):
        self.btn_collect.config(state="disabled", bg="#9CA3AF")
        self.log("크롤링 시작 중...")
        
        threading.Thread(target=self.run_scraping_process, daemon=True).start()

    def run_scraping_process(self):
        try:
            results = scrape_all_sites(log_callback=self.log)
            self.log("엑셀 저장 처리 중...")
            file_name, df_total = export_to_excel(results)
            
            self.current_excel_path = os.path.abspath(file_name)
            self.df_data = df_total
            
            self.root.after(0, self.update_table_view)
            self.root.after(0, lambda: messagebox.showinfo("수집 완료", f"성공적으로 공고 수집이 완료되었습니다!\n\n저장 파일: {file_name}\n총 수집 건수: {len(df_total)}건"))
            self.log(f"수집 완료 (총 {len(df_total)}건)")
        except Exception as e:
            self.log(f"오류 발생: {e}")
            self.root.after(0, lambda: messagebox.showerror("오류", f"수집 과정 중 오류 발생:\n{e}"))
        finally:
            self.root.after(0, lambda: self.btn_collect.config(state="normal", bg="#2563EB"))

    def update_table_view(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        if self.df_data.empty:
            return

        for idx, row in self.df_data.iterrows():
            tag = "even" if idx % 2 == 0 else "odd"
            self.tree.insert("", "end", values=(
                row.get('지자체명', ''),
                row.get('공고제목', ''),
                row.get('매칭키워드', ''),
                row.get('등록일', ''),
                row.get('마감일', ''),
                row.get('링크', '')
            ), tags=(tag,))
            
        self.tree.tag_configure("even", background="#FFFFFF")
        self.tree.tag_configure("odd", background="#F9FAFB")

    def filter_table(self, *args):
        query = self.search_var.get().strip().lower()
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        if self.df_data.empty:
            return
            
        for idx, row in self.df_data.iterrows():
            row_text = f"{row.get('지자체명','')} {row.get('공고제목','')} {row.get('매칭키워드','')} {row.get('등록일','')} {row.get('마감일','')}".lower()
            if not query or query in row_text:
                tag = "even" if idx % 2 == 0 else "odd"
                self.tree.insert("", "end", values=(
                    row.get('지자체명', ''),
                    row.get('공고제목', ''),
                    row.get('매칭키워드', ''),
                    row.get('등록일', ''),
                    row.get('마감일', ''),
                    row.get('링크', '')
                ), tags=(tag,))

    def on_row_double_click(self, event):
        selected = self.tree.selection()
        if not selected:
            return
        item_values = self.tree.item(selected[0], "values")
        if item_values and len(item_values) >= 6:
            link = item_values[5]
            if link.startswith("http"):
                webbrowser.open(link)

    def open_excel_file(self):
        if self.current_excel_path and os.path.exists(self.current_excel_path):
            os.startfile(self.current_excel_path)
        else:
            # Find any recent excel
            files = [f for f in os.listdir('.') if f.startswith('타기관벤치마킹') and f.endswith('.xlsx')]
            if files:
                latest_file = sorted(files, reverse=True)[0]
                os.startfile(os.path.abspath(latest_file))
            else:
                messagebox.showwarning("알림", "생성된 엑셀 파일이 없습니다. 먼저 '공고 실시간 수집 시작' 버튼을 눌러주세요.")

    def load_latest_excel_if_exists(self):
        files = [f for f in os.listdir('.') if f.startswith('타기관벤치마킹') and f.endswith('.xlsx')]
        if files:
            latest_file = sorted(files, reverse=True)[0]
            try:
                self.current_excel_path = os.path.abspath(latest_file)
                self.df_data = pd.read_excel(latest_file, sheet_name='통합비교표')
                self.update_table_view()
                self.log(f"최근 데이터 로드 완료 ({latest_file}, {len(self.df_data)}건)")
            except Exception:
                pass

if __name__ == '__main__':
    root = tk.Tk()
    app = GovernmentNoticeApp(root)
    root.mainloop()
